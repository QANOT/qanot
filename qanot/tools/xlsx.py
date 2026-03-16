"""Excel (.xlsx) spreadsheet tools — create, read, edit."""

from __future__ import annotations

import json
import logging
from pathlib import Path

from qanot.agent import ToolRegistry

logger = logging.getLogger(__name__)


def _fill_sheet(ws, headers: list, rows: list, title: str = "") -> None:
    """Fill a worksheet with headers and rows."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    start_row = 1

    # Title row
    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        start_row = 3

    # Header row
    if headers:
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=str(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = max(len(str(header)) + 4, 12)
        start_row += 1

    # Data rows
    if rows:
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row_idx, row_data in enumerate(rows, start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                # Auto-detect numbers
                if isinstance(value, str):
                    try:
                        cell.value = float(value.replace(",", "").replace(" ", ""))
                        cell.number_format = "#,##0"
                    except (ValueError, AttributeError):
                        pass


def register_xlsx_tools(registry: ToolRegistry, workspace_dir: str) -> None:
    """Register Excel spreadsheet tools."""

    # ── create_xlsx ──
    async def create_xlsx(params: dict) -> str:
        """Create an Excel spreadsheet."""
        try:
            from openpyxl import Workbook
        except ImportError:
            return json.dumps({"error": "openpyxl kutubxonasi o'rnatilmagan. pip install openpyxl"})

        filename = params.get("filename", "spreadsheet.xlsx")
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        sheets = params.get("sheets")
        headers = params.get("headers")
        rows = params.get("rows")
        title = params.get("title", "")

        wb = Workbook()

        if sheets and isinstance(sheets, list):
            for idx, sheet_data in enumerate(sheets):
                if idx == 0:
                    ws = wb.active
                else:
                    ws = wb.create_sheet()
                ws.title = sheet_data.get("name", f"Sheet{idx + 1}")
                _fill_sheet(
                    ws,
                    sheet_data.get("headers", []),
                    sheet_data.get("rows", []),
                    sheet_data.get("title", ""),
                )
        else:
            ws = wb.active
            ws.title = title or "Sheet1"
            _fill_sheet(ws, headers or [], rows or [], title)

        filepath = Path(workspace_dir) / filename
        wb.save(str(filepath))

        return json.dumps({
            "status": "ok",
            "file": str(filepath),
            "filename": filename,
            "message": f"{filename} yaratildi",
        })

    # ── read_xlsx ──
    async def read_xlsx(params: dict) -> str:
        """Read data from Excel spreadsheet."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return json.dumps({"error": "openpyxl kutubxonasi o'rnatilmagan"})

        filepath = params.get("file", "")
        if not filepath:
            return json.dumps({"error": "file parametri kerak"})

        path = Path(filepath) if Path(filepath).is_absolute() else Path(workspace_dir) / filepath
        if not path.exists():
            return json.dumps({"error": f"Fayl topilmadi: {filepath}"})

        wb = load_workbook(str(path), read_only=True, data_only=True)
        sheet_name = params.get("sheet")  # None = active sheet
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])

        sheet_names = wb.sheetnames
        sheet_title = ws.title
        wb.close()

        return json.dumps({
            "sheet": sheet_title,
            "sheets": sheet_names,
            "rows": rows[:500],
            "total_rows": len(rows),
            "truncated": len(rows) > 500,
        }, ensure_ascii=False)

    # ── edit_xlsx ──
    async def edit_xlsx(params: dict) -> str:
        """Edit an existing Excel spreadsheet — append rows, update cells, add sheet."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return json.dumps({"error": "openpyxl kutubxonasi o'rnatilmagan"})

        filepath = params.get("file", "")
        if not filepath:
            return json.dumps({"error": "file parametri kerak"})

        path = Path(filepath) if Path(filepath).is_absolute() else Path(workspace_dir) / filepath
        if not path.exists():
            return json.dumps({"error": f"Fayl topilmadi: {filepath}"})

        wb = load_workbook(str(path))
        sheet_name = params.get("sheet")
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        action = params.get("action", "append_rows")

        if action == "append_rows":
            rows = params.get("rows", [])
            for row_data in rows:
                ws.append(row_data)

        elif action == "update_cell":
            cell = params.get("cell", "")  # e.g. "A1", "B5"
            value = params.get("value", "")
            if cell:
                ws[cell] = value

        elif action == "add_sheet":
            new_name = params.get("new_sheet_name", "Sheet")
            ws_new = wb.create_sheet(title=new_name)
            headers = params.get("headers", [])
            if headers:
                ws_new.append(headers)

        wb.save(str(path))
        wb.close()

        return json.dumps({"status": "ok", "message": f"{path.name} yangilandi", "action": action})

    # Register tools
    registry.register(
        name="create_xlsx",
        description=(
            "Excel (.xlsx) jadval yaratish. Hisobot, ro'yxat, statistika uchun. "
            "headers va rows bering — chiroyli formatlanadi. "
            "Ko'p sahifali Excel uchun sheets parametrini ishlating."
        ),
        parameters={
            "type": "object",
            "required": ["filename"],
            "properties": {
                "filename": {"type": "string", "description": "Fayl nomi (masalan: hisobot.xlsx)"},
                "title": {"type": "string", "description": "Sarlavha (birinchi qator)"},
                "headers": {
                    "type": "array",
                    "description": "Ustun sarlavhalari: [\"Sana\", \"Summa\", \"Mijoz\"]",
                    "items": {"type": "string"},
                },
                "rows": {
                    "type": "array",
                    "description": "Ma'lumot qatorlari: [[\"2026-03-16\", 5000000, \"Sardor\"]]",
                    "items": {"type": "array"},
                },
                "sheets": {
                    "type": "array",
                    "description": "Ko'p sahifa: [{name, title, headers, rows}]",
                    "items": {"type": "object"},
                },
            },
        },
        handler=create_xlsx,
    )

    registry.register(
        name="read_xlsx",
        description="Excel (.xlsx) fayldan ma'lumot o'qish. Barcha qatorlar va ustunlarni qaytaradi.",
        parameters={
            "type": "object",
            "required": ["file"],
            "properties": {
                "file": {"type": "string", "description": "Fayl nomi yoki to'liq path"},
                "sheet": {"type": "string", "description": "Sahifa nomi (ixtiyoriy, standart — faol sahifa)"},
            },
        },
        handler=read_xlsx,
    )

    registry.register(
        name="edit_xlsx",
        description=(
            "Excel (.xlsx) faylni tahrirlash. Qator qo'shish (append_rows), "
            "katak yangilash (update_cell), yangi sahifa (add_sheet)."
        ),
        parameters={
            "type": "object",
            "required": ["file"],
            "properties": {
                "file": {"type": "string", "description": "Fayl nomi yoki to'liq path"},
                "sheet": {"type": "string", "description": "Sahifa nomi (ixtiyoriy)"},
                "action": {
                    "type": "string",
                    "enum": ["append_rows", "update_cell", "add_sheet"],
                    "description": "Harakat turi",
                },
                "rows": {
                    "type": "array",
                    "description": "Qo'shiladigan qatorlar (append_rows uchun)",
                    "items": {"type": "array"},
                },
                "cell": {"type": "string", "description": "Katak manzili, masalan A1, B5 (update_cell uchun)"},
                "value": {"type": "string", "description": "Yangi qiymat (update_cell uchun)"},
                "new_sheet_name": {"type": "string", "description": "Yangi sahifa nomi (add_sheet uchun)"},
                "headers": {
                    "type": "array",
                    "description": "Ustun sarlavhalari (add_sheet uchun)",
                    "items": {"type": "string"},
                },
            },
        },
        handler=edit_xlsx,
    )

    logger.info("XLSX tools registered: create_xlsx, read_xlsx, edit_xlsx")
