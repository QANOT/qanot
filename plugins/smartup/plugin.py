"""Smartup ERP plugin — katalog, tannarx, mijozlar, narx ro'yxati (Excel).

Smartup (biruni framework) integration API uses HTTP Basic auth — the same web
login/password works directly. All data tools here are READ-ONLY (`$export`).
Heavy + rate-limited endpoints are cached locally (see cache.py); `smartup_sync`
refreshes the cache and the other tools read from it.
"""

from __future__ import annotations

import asyncio
import base64
import json
import logging
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import aiohttp

from qanot.plugins.base import Plugin, ToolDef
from plugins.smartup.cache import SmartupCache

logger = logging.getLogger(__name__)

_DIR = Path(__file__).parent
TOOLS_MD = (_DIR / "TOOLS.md").read_text(encoding="utf-8") if (_DIR / "TOOLS.md").exists() else ""
SOUL_APPEND = (_DIR / "SOUL_APPEND.md").read_text(encoding="utf-8") if (_DIR / "SOUL_APPEND.md").exists() else ""

DEFAULT_BASE_URL = "https://smartup.online"

CURRENCY = {"840": "USD", "860": "UZS", "978": "EUR", "643": "RUB"}

# Export endpoint paths (relative to base_url)
EP_INVENTORY = "/b/anor/mxsx/mr/inventory$export"
EP_PRODUCT_GROUP = "/b/anor/mxsx/mr/product_group$export"
EP_ORDER = "/b/trade/txs/tdeal/order$export"          # sale docs → sale prices
EP_PRICE_TYPE = "/b/anor/api/v2/mkr/price_type$export"
EP_LEGAL_PERSON = "/b/anor/mxsx/mr/legal_person$export"
EP_NATURAL_PERSON = "/b/anor/mxsx/mr/natural_person$export"
EP_BALANCE = "/b/anor/mxsx/mkw/balance$export"


class SmartupClient:
    """HTTP client for Smartup integration API (Basic auth)."""

    def __init__(self, base_url: str, login: str, password: str, project_code: str = "trade"):
        self.base_url = base_url.rstrip("/")
        creds = base64.b64encode(f"{login}:{password}".encode()).decode()
        self._auth = f"Basic {creds}"
        self.project_code = project_code
        self._session: aiohttp.ClientSession | None = None
        self._sema = asyncio.Semaphore(2)  # endpoints are throttled — stay gentle

    async def _get_session(self) -> aiohttp.ClientSession:
        if self._session is None or self._session.closed:
            self._session = aiohttp.ClientSession(
                timeout=aiohttp.ClientTimeout(total=180)
            )
        return self._session

    async def export(self, path: str, body: dict | None = None,
                     extra_headers: dict | None = None) -> Any:
        """POST a `$export` request, return parsed JSON."""
        async with self._sema:
            session = await self._get_session()
            url = f"{self.base_url}{path}"
            headers = {
                "Authorization": self._auth,
                "Content-Type": "application/json",
                "Accept": "application/json",
                "project_code": self.project_code,
            }
            if extra_headers:
                headers.update(extra_headers)
            async with session.post(url, data=json.dumps(body or {}), headers=headers) as resp:
                text = await resp.text()
                if resp.status == 401:
                    raise RuntimeError("Smartup autentifikatsiya xatosi — login/parol noto'g'ri")
                if resp.status == 403:
                    raise RuntimeError("Ruxsat etilmagan (403)")
                if resp.status == 429:
                    raise RuntimeError("Smartup so'rov limiti — birozdan keyin urinib ko'ring (429)")
                if resp.status >= 500:
                    raise RuntimeError(f"Smartup server xatosi ({resp.status})")
                try:
                    data = json.loads(text)
                except json.JSONDecodeError:
                    raise RuntimeError(f"Smartup javobi JSON emas ({resp.status}): {text[:160]}")
                if isinstance(data, dict) and data.get("error_code"):
                    raise RuntimeError(f"{data.get('error_code')}: {data.get('message', 'xato')}")
                return data

    async def test_auth(self) -> bool:
        try:
            await self.export(EP_PRICE_TYPE, {})
            return True
        except Exception as e:
            logger.error("[smartup] Auth test failed: %s", e)
            return False

    async def close(self) -> None:
        if self._session and not self._session.closed:
            await self._session.close()


def _ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _parse_dt(s: str | None) -> datetime | None:
    if not s:
        return None
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y"):
        try:
            return datetime.strptime(s.strip(), fmt)
        except (ValueError, AttributeError):
            continue
    return None


def _to_float(v: Any) -> float | None:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


class QanotPlugin(Plugin):
    """Smartup ERP — katalog, tannarx, mijozlar, narx ro'yxati."""

    name = "smartup"
    description = "Smartup ERP — tovarlar, tannarx, mijozlar va mijozlarga narx ro'yxati (Excel)"
    tools_md = TOOLS_MD
    soul_append = SOUL_APPEND

    def __init__(self):
        self.client: SmartupClient | None = None
        self.cache: SmartupCache | None = None
        self._workspace_dir: str = ""

    async def setup(self, config: dict) -> None:
        login = config.get("login", "")
        password = config.get("password", "")
        if not all([login, password]):
            logger.warning("[smartup] Missing config (login, password)")
            return
        base_url = config.get("base_url", DEFAULT_BASE_URL)
        project_code = config.get("project_code", "trade")
        self._workspace_dir = config.get("workspace_dir", "")
        self.client = SmartupClient(base_url, login, password, project_code)
        if not await self.client.test_auth():
            logger.error("[smartup] Auth failed — plugin disabled")
            await self.client.close()
            self.client = None
            return
        db_path = Path(self._workspace_dir or ".") / "smartup" / "cache.db"
        self.cache = SmartupCache(db_path)
        logger.info("[smartup] Plugin ready (base=%s)", base_url)

    async def teardown(self) -> None:
        if self.client:
            await self.client.close()

    def get_tools(self) -> list[ToolDef]:
        if not self.client or not self.cache:
            return []
        tools = self._build_tools()
        logger.info("[smartup] %d tools registered", len(tools))
        return tools

    # ── helpers ───────────────────────────────────────────────
    def _ok(self, data: Any) -> str:
        return json.dumps(data, indent=2, ensure_ascii=False, default=str)

    def _err(self, msg: str) -> str:
        return json.dumps({"error": msg}, ensure_ascii=False)

    @staticmethod
    def _cur(code: str | None) -> str:
        return CURRENCY.get(str(code), str(code) if code else "")

    def _compute_sale_prices(self, orders: list[dict],
                             pt_names: dict[str, str]) -> list[dict]:
        """Aggregate sale orders into the latest sale price per (product, price_type).

        Each order line carries the sold ``product_price`` and ``price_type_code``;
        the order header carries ``currency_code``. We keep, per product and price
        type, the price from the most recent order (by ``deal_time``).
        """
        # (product_code, pt_code) -> (dt, price, currency, pt_name)
        per: dict[tuple[str, str], tuple[datetime, float, str, str]] = {}
        for o in orders:
            dt = _parse_dt(o.get("deal_time")) or datetime.min
            cur = str(o.get("currency_code") or "")
            for ln in o.get("order_products") or []:
                code = ln.get("product_code")
                price = _to_float(ln.get("product_price"))
                if not code or price is None or price <= 0:
                    continue
                ptc = str(ln.get("price_type_code") or "")
                key = (code, ptc)
                prev = per.get(key)
                if prev is None or dt >= prev[0]:
                    per[key] = (dt, price, cur, pt_names.get(ptc, ptc or "—"))

        rows = []
        for (code, ptc), (dt, price, cur, pt_name) in per.items():
            rows.append({
                "product_code": code,
                "price_type_code": ptc,
                "price_type_name": pt_name,
                "price": round(price, 4),
                "currency": cur,
                "date": dt.strftime("%d.%m.%Y") if dt != datetime.min else "",
            })
        return rows

    @staticmethod
    def _date_windows(days: int, max_span: int = 31) -> list[tuple[str, str]]:
        """Split [today-days, today] into <=max_span-day dd.mm.yyyy windows."""
        end = datetime.now()
        start = end - timedelta(days=days)
        windows = []
        cur = start
        while cur < end:
            nxt = min(cur + timedelta(days=max_span), end)
            windows.append((cur.strftime("%d.%m.%Y"), nxt.strftime("%d.%m.%Y")))
            cur = nxt
        return windows

    # ── tools ─────────────────────────────────────────────────
    def _build_tools(self) -> list[ToolDef]:
        c = self.client
        cache = self.cache
        assert c is not None and cache is not None
        tools: list[ToolDef] = []

        # 1 ── sync
        async def sync_handler(p: dict) -> str:
            try:
                price_days = int(p.get("price_days", 90))
                what = p.get("what", "all")
                result: dict[str, Any] = {}

                if what in ("all", "catalog"):
                    groups_raw = (await c.export(EP_PRODUCT_GROUP, {})).get("product_group", [])
                    cat_rows = []
                    for g in groups_raw:
                        for t in g.get("product_group_types") or []:
                            if t.get("product_type_id"):
                                cat_rows.append({
                                    "type_id": t["product_type_id"],
                                    "group_code": g.get("code"),
                                    "group_name": g.get("name"),
                                    "name": t.get("name"),
                                    "code": t.get("code"),
                                })
                    result["categories"] = await cache.replace_categories(cat_rows)

                    inv = (await c.export(EP_INVENTORY, {})).get("inventory", [])
                    result["products"] = await cache.replace_products(inv)
                    await cache.set_meta("last_catalog_sync", _ts())

                if what in ("all", "prices"):
                    # price_type_code -> name (for human-readable price types)
                    pt_names: dict[str, str] = {}
                    try:
                        for t in (await c.export(EP_PRICE_TYPE, {})).get("data", []):
                            code = str(t.get("code") or "")
                            if code:
                                pt_names[code] = t.get("name") or code
                    except Exception as e:  # noqa: BLE001 — names are a nicety, not required
                        logger.warning("[smartup] price_type names unavailable: %s", e)

                    orders: dict[str, dict] = {}
                    for begin, end in self._date_windows(price_days):
                        chunk = (await c.export(EP_ORDER,
                                                {"begin_date": begin, "end_date": end})).get("order", [])
                        for d in chunk:
                            orders[d.get("deal_id") or f"{begin}-{len(orders)}"] = d
                    price_rows = self._compute_sale_prices(list(orders.values()), pt_names)
                    result["sale_prices"] = await cache.replace_sale_prices(price_rows)
                    result["order_docs"] = len(orders)
                    result["price_days"] = price_days
                    await cache.set_meta("last_price_sync", _ts())

                if what in ("all", "customers"):
                    custs: list[dict] = []
                    for ep, kind in ((EP_LEGAL_PERSON, "legal"), (EP_NATURAL_PERSON, "natural")):
                        try:
                            data = await c.export(ep, {})
                            key = "legal_person" if kind == "legal" else "natural_person"
                            for r in data.get(key, []):
                                r["_kind"] = kind
                                custs.append(r)
                        except Exception as e:  # noqa: BLE001 — one kind missing shouldn't fail sync
                            logger.warning("[smartup] %s sync failed: %s", kind, e)
                    if custs:
                        result["customers"] = await cache.replace_customers(custs)

                result["synced_at"] = _ts()
                return self._ok(result)
            except Exception as e:
                return self._err(str(e))

        tools.append(ToolDef(
            name="smartup_sync",
            description=("Smartup'dan ma'lumotni lokal keshga yuklab oladi (tovarlar, "
                         "kategoriyalar, sotuv narxlari, mijozlar). Boshqa smartup_* "
                         "tool'lar shu keshdan o'qiydi. Birinchi marta va ma'lumot "
                         "eskirganda chaqiring. Og'ir amal — kunda bir necha marta yetadi."),
            parameters={"type": "object", "properties": {
                "what": {"type": "string", "enum": ["all", "catalog", "prices", "customers"],
                         "description": "Nimani yangilash (default: all)"},
                "price_days": {"type": "number",
                               "description": "Sotuv narxlari uchun necha kun orqaga qarash (default 90)"},
            }},
            handler=sync_handler,
        ))

        # 2 ── cache status
        async def status_handler(p: dict) -> str:
            try:
                return self._ok(await cache.stats())
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            name="smartup_status",
            description="Lokal kesh holati — nechta tovar/kategoriya/narx/mijoz va oxirgi yangilanish vaqti.",
            parameters={"type": "object", "properties": {}},
            handler=status_handler,
        ))

        # 3 ── categories
        async def categories_handler(p: dict) -> str:
            try:
                cats = await cache.list_categories()
                if not cats:
                    return self._ok({"categories": [], "hint": "Kesh bo'sh — avval smartup_sync chaqiring."})
                return self._ok({"categories": cats, "total": len(cats)})
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            name="smartup_categories",
            description="Tovar kategoriyalari ro'yxati (har birida tovar soni). Keshdan o'qiydi.",
            parameters={"type": "object", "properties": {}},
            handler=categories_handler,
        ))

        # 4 ── price types
        async def price_types_handler(p: dict) -> str:
            try:
                pts = await cache.list_price_types()
                if not pts:
                    return self._ok({"price_types": [],
                                     "hint": "Kesh bo'sh — avval smartup_sync chaqiring."})
                out = [{"price_type_code": t["price_type_code"],
                        "name": t["price_type_name"],
                        "currency": self._cur(t["currency"]),
                        "product_count": t["product_count"]} for t in pts]
                return self._ok({"price_types": out})
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            name="smartup_price_types",
            description=("Mavjud sotuv narx-turlari (masalan ulgurji USD, chakana UZS) va "
                         "har biriga narxi bor tovarlar soni. Narx ro'yxati uchun qaysi "
                         "narx-turini ishlatishni tanlashda yordam beradi."),
            parameters={"type": "object", "properties": {}},
            handler=price_types_handler,
        ))

        # 5 ── search products
        async def search_handler(p: dict) -> str:
            try:
                rows = await cache.search_products(
                    query=str(p.get("query", "")),
                    category=str(p.get("category", "")),
                    price_type_code=str(p.get("price_type_code", "")),
                    limit=int(p.get("limit", 50)),
                )
                out = []
                for r in rows:
                    prices = [{
                        "price_type": pr["price_type_name"],
                        "price_type_code": pr["price_type_code"],
                        "price": pr["price"],
                        "currency": self._cur(pr["currency"]),
                        "date": pr["date"],
                    } for pr in r.get("prices", [])]
                    out.append({
                        "code": r["code"], "name": r["name"],
                        "article_code": r["article_code"],
                        "category": r["category_names"],
                        "measure": r["measure_code"],
                        "prices": prices,
                    })
                return self._ok({"products": out, "total": len(out),
                                 "note": "Narxsiz tovarlar so'nggi sotuvlar oralig'ida sotilmagan."})
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            name="smartup_search_products",
            description=("Tovarlarni nom/artikul/shtrix-kod yoki kategoriya bo'yicha keshdan qidiradi. "
                         "Har tovarda sotuv narxlari (narx-turi bo'yicha) qaytadi."),
            parameters={"type": "object", "properties": {
                "query": {"type": "string", "description": "Nom, artikul, kod yoki shtrix-kod (qism)"},
                "category": {"type": "string", "description": "Kategoriya nomi (qism)"},
                "price_type_code": {"type": "string", "description": "Faqat shu narx-turi (smartup_price_types'dan)"},
                "limit": {"type": "number", "description": "Maksimal natija (default 50)"},
            }},
            handler=search_handler,
        ))

        # 6 ── pricelist (Excel)
        async def pricelist_handler(p: dict) -> str:
            try:
                query = str(p.get("query", ""))
                category = str(p.get("category", ""))
                if not query and not category:
                    return self._err("query yoki category — kamida bittasi kerak (qaysi tovarlar?).")
                price_type_code = str(p.get("price_type_code", ""))
                markup = _to_float(p.get("markup_percent")) or 0.0
                round_to = _to_float(p.get("round_to")) or 0.0
                limit = int(p.get("limit", 2000))
                title = str(p.get("title", "")) or (category or query or "Narx ro'yxati")

                rows = await cache.search_products(
                    query=query, category=category,
                    price_type_code=price_type_code, limit=limit)
                if not rows:
                    return self._ok({"generated": False,
                                     "hint": "Mos tovar topilmadi. Kesh bo'sh bo'lsa avval smartup_sync."})

                items, without_price = [], 0
                cur_breakdown: dict[str, int] = {}
                for r in rows:
                    prices = r.get("prices", [])
                    if not prices:
                        without_price += 1
                        continue
                    # pick chosen price type, else the first available
                    chosen = prices[0]
                    base = chosen["price"]
                    sale = base * (1 + markup / 100.0)
                    if round_to and round_to > 0:
                        sale = round(sale / round_to) * round_to
                    cur = self._cur(chosen["currency"])
                    cur_breakdown[cur] = cur_breakdown.get(cur, 0) + 1
                    items.append({
                        "code": r["code"], "article": r["article_code"], "name": r["name"],
                        "category": r["category_names"], "measure": r["measure_code"],
                        "price_type": chosen["price_type_name"],
                        "base": round(base, 2), "currency": cur, "date": chosen["date"],
                        "sale": round(sale, 2),
                    })

                if not items:
                    return self._ok({"generated": False, "total_matched": len(rows),
                                     "without_price": without_price,
                                     "hint": ("Topilgan tovarlarda sotuv narxi yo'q. "
                                              "smartup_sync(what='prices', price_days=180) bilan oraliqni kengaytiring.")})

                file_path = await asyncio.to_thread(
                    self._write_pricelist_xlsx, title, markup, items)

                return self._ok({
                    "generated": True,
                    "title": title,
                    "markup_percent": markup,
                    "price_type_code": price_type_code or "(har tovarning birinchi narx-turi)",
                    "round_to": round_to,
                    "total": len(items),
                    "without_price": without_price,
                    "currency_breakdown": cur_breakdown,
                    "preview": items[:15],
                    "file_path": file_path,
                    "format": "xlsx",
                    "advice": "Foydalanuvchiga yuborish uchun: send_file(file_path).",
                })
            except Exception as e:
                return self._err(str(e))

        tools.append(ToolDef(
            name="smartup_pricelist",
            description=("Mijozlarga narx ro'yxatini Excel (.xlsx) qilib tayyorlaydi. "
                         "Tovarlarni query/category bo'yicha tanlaydi, SOTUV narxini oladi "
                         "(narx asl valyutada — USD/UZS). Ixtiyoriy markup_percent ustama "
                         "qo'shadi (0 = sof sotuv narxi). Bir tovarda bir nechta narx-turi "
                         "bo'lsa price_type_code bilan aniqlang (smartup_price_types ko'rsatadi). "
                         "Natijada file_path qaytadi — send_file bilan yuboring."),
            parameters={"type": "object", "properties": {
                "query": {"type": "string", "description": "Tovar nomi/artikul/kod (qism). bambuk kabi so'z nom bo'yicha qidiradi."},
                "category": {"type": "string", "description": "Kategoriya nomi (qism)"},
                "price_type_code": {"type": "string", "description": "Qaysi sotuv narx-turi (masalan ulgurji USD). Bo'sh bo'lsa har tovarning mavjud birinchi narxi."},
                "markup_percent": {"type": "number", "description": "Sotuv narxiga ustama foiz (default 0 = ustamasiz)"},
                "round_to": {"type": "number", "description": "Yakuniy narxni yaxlitlash qadami (masalan 1000). 0 = yaxlitlamaslik"},
                "title": {"type": "string", "description": "Ro'yxat sarlavhasi / fayl nomi"},
                "limit": {"type": "number", "description": "Maksimal tovar (default 2000)"},
            }},
            handler=pricelist_handler,
        ))

        # 6 ── customers
        async def customers_handler(p: dict) -> str:
            try:
                rows = await cache.search_customers(
                    query=str(p.get("query", "")),
                    kind=str(p.get("kind", "")),
                    limit=int(p.get("limit", 50)),
                )
                return self._ok({"customers": rows, "total": len(rows)})
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            name="smartup_customers",
            description="Mijoz/ta'minotchilarni (yuridik va jismoniy shaxslar) keshdan qidiradi.",
            parameters={"type": "object", "properties": {
                "query": {"type": "string", "description": "Ism, telefon, STIR yoki kod (qism)"},
                "kind": {"type": "string", "enum": ["legal", "natural"],
                         "description": "Shaxs turi: legal (yuridik) yoki natural (jismoniy)"},
                "limit": {"type": "number", "description": "Maksimal natija (default 50)"},
            }},
            handler=customers_handler,
        ))

        # 7 ── balance (live, read-only)
        async def balance_handler(p: dict) -> str:
            try:
                filial_code = str(p.get("filial_code", ""))
                if not filial_code:
                    return self._err("filial_code majburiy (tashkilot/filial kodi).")
                begin = str(p.get("begin_date", ""))
                end = str(p.get("end_date", ""))
                if not begin or not end:
                    end_dt = datetime.now()
                    begin = (end_dt - timedelta(days=30)).strftime("%d.%m.%Y")
                    end = end_dt.strftime("%d.%m.%Y")
                wh = str(p.get("warehouse_code", ""))
                body = {
                    "warehouse_codes": [{"warehouse_code": wh}],
                    "filial_code": filial_code,
                    "begin_date": begin, "end_date": end,
                }
                data = await c.export(EP_BALANCE, body)
                return self._ok(data)
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            name="smartup_balance",
            description=("Ombor qoldig'i (jonli so'rov). filial_code va sana oralig'i (31 kundan kam) "
                         "kerak. Sana berilmasa oxirgi 30 kun olinadi."),
            parameters={"type": "object", "required": ["filial_code"], "properties": {
                "filial_code": {"type": "string", "description": "Tashkilot/filial kodi"},
                "warehouse_code": {"type": "string", "description": "Ombor kodi (bo'sh = barchasi)"},
                "begin_date": {"type": "string", "description": "Boshlanish dd.mm.yyyy"},
                "end_date": {"type": "string", "description": "Tugash dd.mm.yyyy (oraliq <31 kun)"},
            }},
            handler=balance_handler,
        ))

        return tools

    def _write_pricelist_xlsx(self, title: str, markup: float,
                              items: list[dict]) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        out_dir = Path(self._workspace_dir or ".") / "generated"
        out_dir.mkdir(parents=True, exist_ok=True)
        safe = "".join(ch if ch.isalnum() else "_" for ch in title)[:40] or "pricelist"
        path = out_dir / f"smartup_pricelist_{safe}_{int(time.time())}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Narxlar"

        subtitle = f"{title}  —  sotuv narxi" + (f" + {markup:g}% ustama" if markup else "")
        ws.append([subtitle])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        ws["A1"].font = Font(bold=True, size=13)

        headers = ["№", "Kod", "Artikul", "Nomi", "Narx turi", "Narx", "Valyuta", "Yakuniy narx"]
        ws.append(headers)
        hdr_fill = PatternFill("solid", fgColor="D9E1F2")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for i, it in enumerate(items, start=1):
            ws.append([i, it["code"], it["article"], it["name"], it["price_type"],
                       it["base"], it["currency"], it["sale"]])

        widths = [5, 16, 14, 50, 22, 12, 9, 14]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w
        for col_idx in (6, 8):
            for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = "#,##0.00"
        ws.freeze_panes = "A3"

        wb.save(path)
        return str(path)
