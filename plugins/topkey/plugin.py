"""TopKey HR + Project Management plugin (28 tools)."""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

from qanot.plugins.base import Plugin, ToolDef
from qanot.secrets import resolve_secret

logger = logging.getLogger(__name__)


def _import_client():
    """Import the HTTP client lazily.

    Plugin's own dir is added to sys.path by the loader before plugin.py is
    executed, but tests that `from plugins.topkey.plugin import QanotPlugin`
    don't go through the loader. Doing the import here covers both paths and
    avoids putting `tk_engine` on sys.path globally at module load time.
    """
    import sys
    plugin_dir = str(Path(__file__).parent)
    if plugin_dir not in sys.path:
        sys.path.insert(0, plugin_dir)
    from tk_engine.client import TopKeyAPIError, TopKeyAuthError, TopKeyClient
    return TopKeyClient, TopKeyAuthError, TopKeyAPIError

_PLUGIN_DIR = Path(__file__).parent
TOOLS_MD = (_PLUGIN_DIR / "TOOLS.md").read_text(encoding="utf-8") if (_PLUGIN_DIR / "TOOLS.md").exists() else ""
SOUL_APPEND = (_PLUGIN_DIR / "SOUL_APPEND.md").read_text(encoding="utf-8") if (_PLUGIN_DIR / "SOUL_APPEND.md").exists() else ""


class QanotPlugin(Plugin):
    """TopKey HR + PM plugin."""

    name = "topkey"
    description = "TopKey HR & Project Management — xodimlar, davomat, ta'tillar, loyihalar, vazifalar"
    tools_md = TOOLS_MD
    soul_append = SOUL_APPEND

    def __init__(self):
        self.client: Any = None
        self._workspace_dir: str = ""

    async def setup(self, config: dict) -> None:
        api_url = config.get("api_url", "")
        email = config.get("email", "")
        # Password may be plain or a SecretRef ({"env": "..."}/{"file": "..."}).
        # The framework only auto-resolves a fixed set of top-level secret
        # fields (api_key, bot_token, ...) — plugin configs are passed through
        # untouched, so the plugin resolves its own SecretRef shapes.
        try:
            password = resolve_secret(config.get("password", ""))
        except Exception as e:
            logger.warning("[topkey] Password resolution failed: %s", e)
            password = ""
        self._workspace_dir = config.get("workspace_dir", "")
        if not api_url or not email or not password:
            logger.warning("[topkey] Missing config (api_url, email, password) — plugin disabled")
            return
        TopKeyClient, TopKeyAuthError, _ = _import_client()
        self.client = TopKeyClient(api_url, email, password)
        try:
            await self.client.login()
            logger.info("[topkey] Logged in successfully")
        except TopKeyAuthError as e:
            logger.error("[topkey] Login failed: %s", e)
            self.client = None
        except Exception as e:
            logger.error("[topkey] Unexpected error during login: %s", e)
            self.client = None

    async def teardown(self) -> None:
        if self.client:
            await self.client.close()

    def get_tools(self) -> list[ToolDef]:
        if not self.client:
            return []
        tools = self._build_tools()
        logger.info("[topkey] %d tools registered", len(tools))
        return tools

    # ── Helpers ───────────────────────────────────────────────

    # Spool threshold for list responses: when a tool returns more than this
    # many rows, write the full result to {workspace}/generated/<file>.xlsx
    # and surface a preview + file_path. Avoids the agent inlining huge
    # payloads + losing iterations on manual pagination (the absmarket bug
    # that ate 25 iterations on Davron's 8,675-row report).
    _SPOOL_THRESHOLD = 100

    def _maybe_spool(self, name: str, items: list, full: dict) -> str:
        """If `items` is large, write Excel and return a reference; else inline."""
        total = len(items)
        if total <= self._SPOOL_THRESHOLD or not items:
            return self._ok({"items": items, "total": total, **{k: v for k, v in full.items() if k not in ("items",)}})

        try:
            import time
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except Exception as e:  # noqa: BLE001 — missing dep degrades to inline
            logger.warning("[topkey] openpyxl unavailable, returning inline: %s", e)
            return self._ok({"items": items[:self._SPOOL_THRESHOLD], "total": total, "truncated": True})

        out_dir = Path(self._workspace_dir or ".") / "generated"
        out_dir.mkdir(parents=True, exist_ok=True)
        ts = int(time.time())
        path = out_dir / f"topkey_{name}_{ts}.xlsx"

        # Flatten one level: nested dicts -> JSON strings to keep cells single-valued
        cols: list[str] = []
        seen = set()
        for r in items:
            if isinstance(r, dict):
                for k in r.keys():
                    if k not in seen:
                        seen.add(k)
                        cols.append(k)
        wb = Workbook()
        ws = wb.active
        ws.title = name[:31] or "topkey"
        ws.append(cols)
        for r in items:
            row = []
            for c in cols:
                v = r.get(c) if isinstance(r, dict) else None
                if isinstance(v, (dict, list)):
                    v = json.dumps(v, ensure_ascii=False, default=str)
                row.append(v)
            ws.append(row)
        # Auto-width based on first 50 rows
        for idx, col in enumerate(cols, start=1):
            sample = [str(r.get(col, ""))[:80] for r in items[:50] if isinstance(r, dict)]
            width = min(max([len(col)] + [len(s) for s in sample]), 60)
            ws.column_dimensions[get_column_letter(idx)].width = width + 2
        wb.save(path)

        return self._ok({
            "preview": items[:20],
            "total": total,
            "file_path": str(path),
            "format": "xlsx",
            "advice": (
                f"{total} ta yozuv Excel faylga saqlandi. Foydalanuvchiga yuborish uchun: "
                "send_file({file_path}). Qayta {tool_name} chaqirib paginatsiya QILMANG."
            ),
            **{k: v for k, v in full.items() if k != "items"},
        })

    def _ok(self, data: Any) -> str:
        return json.dumps(data, indent=2, ensure_ascii=False, default=str)

    def _err(self, msg: str) -> str:
        return json.dumps({"error": msg}, ensure_ascii=False)

    def _build_tools(self) -> list[ToolDef]:
        c = self.client
        assert c is not None
        tools: list[ToolDef] = []

        # ── HR / EMPLOYEES (5) ───────────────────────────────

        async def list_employees(p: dict) -> str:
            try:
                params: dict[str, Any] = {}
                if "department_id" in p:
                    params["department_id"] = p["department_id"]
                if "designation_id" in p:
                    params["designation_id"] = p["designation_id"]
                if "status" in p:
                    params["status"] = p["status"]
                if "page" in p:
                    params["page"] = p["page"]
                if "per_page" in p:
                    params["per_page"] = p["per_page"]
                result = await c.get("/employee", params or None)
                return self._ok(result)
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_list_employees",
            "TopKey: Xodimlar ro'yxati. Filter: department_id, designation_id, status. Paginatsiya: page, per_page.",
            {"type": "object", "properties": {
                "department_id": {"type": "number"},
                "designation_id": {"type": "number"},
                "status": {"type": "string", "description": "active | deactive"},
                "page": {"type": "number"},
                "per_page": {"type": "number"},
            }},
            list_employees,
        ))

        async def get_employee(p: dict) -> str:
            if "employee_id" not in p:
                return self._err("employee_id is required")
            try:
                return self._ok(await c.get(f"/employee/{int(p['employee_id'])}"))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_get_employee",
            "TopKey: Bitta xodimning to'liq profili (employee_id bo'yicha).",
            {"type": "object", "required": ["employee_id"], "properties": {
                "employee_id": {"type": "number"},
            }},
            get_employee,
        ))

        async def list_departments(_p: dict) -> str:
            try:
                return self._ok(await c.get("/department"))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_list_departments",
            "TopKey: Bo'limlar ro'yxati.",
            {"type": "object", "properties": {}},
            list_departments,
        ))

        async def list_designations(_p: dict) -> str:
            try:
                return self._ok(await c.get("/designation"))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_list_designations",
            "TopKey: Lavozimlar ro'yxati.",
            {"type": "object", "properties": {}},
            list_designations,
        ))

        async def get_user_profile(_p: dict) -> str:
            try:
                return self._ok(await c.get("/auth/me"))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_get_user_profile",
            "TopKey: Joriy autentifikatsiya qilingan foydalanuvchi profili (\"men kim?\" savollari uchun).",
            {"type": "object", "properties": {}},
            get_user_profile,
        ))

        # ── HR / ATTENDANCE (5) ──────────────────────────────

        async def get_today_attendance(p: dict) -> str:
            try:
                params: dict[str, Any] = {}
                if "date" in p:
                    params["date"] = p["date"]
                # /mobile/attendance/all is the admin "today across employees"
                # view; falls back to /attendance/today if `date` is absent.
                path = "/mobile/attendance/all" if params else "/attendance/today"
                return self._ok(await c.get(path, params or None))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_get_today_attendance",
            "TopKey: Bugun (yoki berilgan sanada) check-in qilgan barcha xodimlar (admin ko'rinishi).",
            {"type": "object", "properties": {
                "date": {"type": "string", "description": "YYYY-MM-DD; bo'sh bo'lsa bugungi sana"},
            }},
            get_today_attendance,
        ))

        async def get_user_attendance(p: dict) -> str:
            if "user_id" not in p:
                return self._err("user_id is required")
            try:
                params: dict[str, Any] = {"user_id": p["user_id"]}
                # employeeHistory uses year+month; allow both granularities.
                if "year" in p:
                    params["year"] = p["year"]
                if "month" in p:
                    params["month"] = p["month"]
                if "from_date" in p:
                    params["from_date"] = p["from_date"]
                if "to_date" in p:
                    params["to_date"] = p["to_date"]
                return self._ok(await c.get("/mobile/attendance/employee-history", params))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_get_user_attendance",
            "TopKey: Bitta xodim davomat tarixi (year+month yoki from_date+to_date oralig'ida).",
            {"type": "object", "required": ["user_id"], "properties": {
                "user_id": {"type": "number"},
                "year": {"type": "number"},
                "month": {"type": "number"},
                "from_date": {"type": "string"},
                "to_date": {"type": "string"},
            }},
            get_user_attendance,
        ))

        # Shifts are reused across attendance queries; fetch once and cache
        # in-memory for the lifetime of the plugin (re-fetch is cheap if a
        # shift is added — but rare). Keys: id -> {start_time, late_mark_duration}.
        _shifts_cache: dict[int, dict] = {}

        async def _shifts() -> dict[int, dict]:
            if _shifts_cache:
                return _shifts_cache
            try:
                data = await c.get("/mobile/shift/list")
                items = []
                if isinstance(data, dict):
                    items = data.get("data") or []
                elif isinstance(data, list):
                    items = data
                for s in items:
                    if isinstance(s, dict) and "id" in s:
                        _shifts_cache[int(s["id"])] = {
                            "name": s.get("name", ""),
                            "start_time": s.get("start_time", "00:00:00"),
                            "end_time": s.get("end_time", "00:00:00"),
                            "late_mark_duration": int(s.get("late_mark_duration", 0) or 0),
                        }
            except Exception as e:  # noqa: BLE001 — shift fetch failure shouldn't block whole tool
                logger.warning("[topkey] shift cache fetch failed: %s", e)
            return _shifts_cache

        async def _attendance_rows(date: str | None) -> list[dict]:
            params = {"date": date} if date else {}
            data = await c.get("/mobile/attendance/all", params or None)
            rows: list = []
            if isinstance(data, dict):
                inner = data.get("data")
                if isinstance(inner, list):
                    rows = inner
                elif isinstance(inner, dict) and isinstance(inner.get("data"), list):
                    rows = inner["data"]
            elif isinstance(data, list):
                rows = data
            return [r for r in rows if isinstance(r, dict)]

        def _classify_attendance(rows: list[dict], shifts: dict[int, dict]) -> dict:
            """Return classification: present_on_time, late, absent, on_leave.

            Late is computed against the shift's start_time + late_mark_duration
            (per-shift configurable, e.g. 15 min grace). Without a real shift
            entry we fall back to 09:00 + 15 min for safety.
            """
            from datetime import datetime, timedelta
            on_time: list[dict] = []
            late: list[dict] = []
            absent: list[dict] = []
            on_leave: list[dict] = []

            for r in rows:
                if str(r.get("is_on_leave", "0")) == "1":
                    on_leave.append(r)
                    continue
                check_in = r.get("check_in", "-")
                if not check_in or check_in == "-":
                    absent.append(r)
                    continue
                try:
                    ci_dt = datetime.fromisoformat(str(check_in).replace("Z", "+00:00").replace(" ", "T"))
                except (ValueError, TypeError):
                    on_time.append(r)
                    continue
                shift = shifts.get(int(r.get("shift_id", 0) or 0)) if shifts else None
                start_str = (shift or {}).get("start_time") or "09:00:00"
                grace_min = int((shift or {}).get("late_mark_duration") or 15)
                # Build the threshold datetime using the check-in's date and
                # the shift's start_time. Both are in Asia/Tashkent semantics.
                try:
                    h, m, s = (int(x) for x in start_str.split(":"))
                except (ValueError, TypeError):
                    h, m, s = 9, 0, 0
                threshold = ci_dt.replace(hour=h, minute=m, second=s, microsecond=0) + timedelta(minutes=grace_min)
                # Special case: shift is "Dam olish kuni" (00:00-00:00) — skip late check.
                if start_str == "00:00:00":
                    on_time.append(r)
                    continue
                if ci_dt > threshold:
                    minutes_late = int((ci_dt - threshold).total_seconds() // 60) + grace_min
                    enriched = {**r, "minutes_late": minutes_late, "shift_name": (shift or {}).get("name", "?")}
                    late.append(enriched)
                else:
                    on_time.append(r)
            return {"on_time": on_time, "late": late, "absent": absent, "on_leave": on_leave}

        async def get_team_summary(p: dict) -> str:
            """Aggregate /mobile/attendance/all using real shift schedules."""
            try:
                date = p.get("date")
                shifts = await _shifts()
                rows = await _attendance_rows(date)
                cls = _classify_attendance(rows, shifts)
                return self._ok({
                    "date": date or "today",
                    "total": len(rows),
                    "present_on_time": len(cls["on_time"]),
                    "present_late": len(cls["late"]),
                    "absent": len(cls["absent"]),
                    "on_leave": len(cls["on_leave"]),
                })
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_get_team_summary",
            (
                "TopKey: Kunlik davomat sarhisobi. 'Kech' status har shift'ning "
                "real start_time + late_mark_duration (smena ta'rifidan) "
                "asosida hisoblanadi."
            ),
            {"type": "object", "properties": {
                "date": {"type": "string", "description": "YYYY-MM-DD; bo'sh — bugun"},
            }},
            get_team_summary,
        ))

        async def get_late_arrivals(p: dict) -> str:
            """Return employees who arrived later than their shift start +
            late_mark_duration. Shift schedules are pulled from /shift/list."""
            try:
                date = p.get("date")
                shifts = await _shifts()
                rows = await _attendance_rows(date)
                cls = _classify_attendance(rows, shifts)
                late = cls["late"]
                # Sort by minutes_late desc so the agent sees the most extreme first
                late.sort(key=lambda r: r.get("minutes_late", 0), reverse=True)
                return self._ok({
                    "date": date or "today",
                    "count": len(late),
                    "employees": [
                        {
                            "user_id": r.get("user_id"),
                            "name": r.get("name"),
                            "check_in": r.get("check_in"),
                            "minutes_late": r.get("minutes_late"),
                            "shift_name": r.get("shift_name"),
                        }
                        for r in late
                    ],
                })
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_get_late_arrivals",
            (
                "TopKey: Berilgan sanada kech kelgan xodimlar (har biri uchun "
                "minutes_late va smena nomi). Shift start_time + "
                "late_mark_duration (smena grace minutes) asosida hisoblanadi."
            ),
            {"type": "object", "properties": {
                "date": {"type": "string", "description": "YYYY-MM-DD; bo'sh — bugun"},
            }},
            get_late_arrivals,
        ))

        async def get_overtime(p: dict) -> str:
            try:
                params: dict[str, Any] = {}
                if "user_id" in p:
                    params["user_id"] = p["user_id"]
                if "from_date" in p:
                    params["from_date"] = p["from_date"]
                if "to_date" in p:
                    params["to_date"] = p["to_date"]
                # Attendance index supports filters; overtime fields surface in
                # individual attendance records (overtime_hours, total_hours).
                return self._ok(await c.get("/attendance", params or None))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_get_overtime",
            "TopKey: Sverxurochniy hisobot — xodim va sana oralig'i bo'yicha (attendance recordlardagi overtime_hours).",
            {"type": "object", "properties": {
                "user_id": {"type": "number"},
                "from_date": {"type": "string"},
                "to_date": {"type": "string"},
            }},
            get_overtime,
        ))

        # ── HR / LEAVE (5) ───────────────────────────────────

        async def list_leave_requests(p: dict) -> str:
            try:
                params: dict[str, Any] = {}
                if "status" in p:
                    params["status"] = p["status"]
                if "user_id" in p:
                    params["user_id"] = p["user_id"]
                if "from_date" in p:
                    params["from_date"] = p["from_date"]
                if "to_date" in p:
                    params["to_date"] = p["to_date"]
                if "page" in p:
                    params["page"] = p["page"]
                if "per_page" in p:
                    params["per_page"] = p["per_page"]
                return self._ok(await c.get("/leave", params or None))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_list_leave_requests",
            "TopKey: Ta'til so'rovlari ro'yxati. Filter: status (pending|approved|rejected), user_id, from_date/to_date.",
            {"type": "object", "properties": {
                "status": {"type": "string"},
                "user_id": {"type": "number"},
                "from_date": {"type": "string"},
                "to_date": {"type": "string"},
                "page": {"type": "number"},
                "per_page": {"type": "number"},
            }},
            list_leave_requests,
        ))

        async def create_leave_request(p: dict) -> str:
            for k in ("user_id", "leave_type_id", "start_date", "end_date"):
                if k not in p:
                    return self._err(f"{k} is required")
            try:
                body = {
                    "user_ids": [int(p["user_id"])],
                    "leave_type_id": int(p["leave_type_id"]),
                    "duration": p.get("duration", "multiple"),
                    "start_date": p["start_date"],
                    "end_date": p["end_date"],
                    "reason": p.get("reason", ""),
                }
                # /mobile/leave/grant is the admin "grant leave" endpoint that
                # accepts the multi-user shape; falls back to /leave/apply for
                # the single-user case.
                return self._ok(await c.post("/mobile/leave/grant", body))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_create_leave_request",
            "TopKey: Admin tomonidan ta'til berish (user_id, leave_type_id, start_date, end_date, reason).",
            {"type": "object", "required": ["user_id", "leave_type_id", "start_date", "end_date"], "properties": {
                "user_id": {"type": "number"},
                "leave_type_id": {"type": "number"},
                "start_date": {"type": "string", "description": "YYYY-MM-DD"},
                "end_date": {"type": "string", "description": "YYYY-MM-DD"},
                "duration": {"type": "string", "description": "single | multiple (default: multiple)"},
                "reason": {"type": "string"},
            }},
            create_leave_request,
        ))

        async def approve_leave(p: dict) -> str:
            if "leave_id" not in p:
                return self._err("leave_id is required")
            try:
                body = {"approve_reason": p.get("approve_reason", p.get("comment", ""))}
                return self._ok(await c.put(f"/mobile/leave/{int(p['leave_id'])}/approve", body))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_approve_leave",
            "TopKey: Ta'til so'rovini tasdiqlash (leave_id, ixtiyoriy approve_reason).",
            {"type": "object", "required": ["leave_id"], "properties": {
                "leave_id": {"type": "number"},
                "approve_reason": {"type": "string"},
            }},
            approve_leave,
        ))

        async def get_leave_balance(p: dict) -> str:
            if "user_id" not in p:
                return self._err("user_id is required")
            try:
                # No dedicated balance endpoint exists; the admin grant flow
                # surfaces remaining days in the user employeeDetail. For now
                # we return the user with leave types so the agent can compute
                # remaining = entitlement - used.
                user = await c.get(f"/employee/{int(p['user_id'])}")
                types = await c.get("/leave-type")
                return self._ok({"user": user, "leave_types": types})
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_get_leave_balance",
            "TopKey: Xodimning qolgan ta'til kunlari (turi bo'yicha).",
            {"type": "object", "required": ["user_id"], "properties": {
                "user_id": {"type": "number"},
            }},
            get_leave_balance,
        ))

        async def list_leave_types(_p: dict) -> str:
            try:
                return self._ok(await c.get("/leave-type"))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_list_leave_types",
            "TopKey: Ta'til turlari ro'yxati.",
            {"type": "object", "properties": {}},
            list_leave_types,
        ))

        # ── WORK / PROJECTS (3) ──────────────────────────────

        async def list_projects(p: dict) -> str:
            try:
                params: dict[str, Any] = {}
                if "status" in p:
                    params["status"] = p["status"]
                if "client_id" in p:
                    params["client_id"] = p["client_id"]
                if "category_id" in p:
                    params["category_id"] = p["category_id"]
                # Auto-walk pages so the agent doesn't loop.
                if p.get("all"):
                    return self._ok(await c.get_all("/project", params or None))
                if "page" in p:
                    params["page"] = p["page"]
                if "per_page" in p:
                    params["per_page"] = p["per_page"]
                return self._ok(await c.get("/project", params or None))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_list_projects",
            "TopKey: Loyihalar ro'yxati. Filter: status, client_id, category_id. all=true bo'lsa avto-paginatsiya (max 5 sahifa / 500 ta).",
            {"type": "object", "properties": {
                "status": {"type": "string"},
                "client_id": {"type": "number"},
                "category_id": {"type": "number"},
                "page": {"type": "number"},
                "per_page": {"type": "number"},
                "all": {"type": "boolean", "description": "barcha sahifalarni o'qish"},
            }},
            list_projects,
        ))

        async def get_project(p: dict) -> str:
            if "project_id" not in p:
                return self._err("project_id is required")
            try:
                return self._ok(await c.get(f"/project/{int(p['project_id'])}"))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_get_project",
            "TopKey: Loyiha tafsilotlari (project_id bo'yicha).",
            {"type": "object", "required": ["project_id"], "properties": {
                "project_id": {"type": "number"},
            }},
            get_project,
        ))

        async def list_project_members(p: dict) -> str:
            if "project_id" not in p:
                return self._err("project_id is required")
            try:
                return self._ok(await c.get(f"/project/{int(p['project_id'])}/members"))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_list_project_members",
            "TopKey: Loyihaga biriktirilgan xodimlar.",
            {"type": "object", "required": ["project_id"], "properties": {
                "project_id": {"type": "number"},
            }},
            list_project_members,
        ))

        # ── WORK / TASKS (5) ─────────────────────────────────

        async def list_tasks(p: dict) -> str:
            """List ALL tasks via Froiden offset-pagination, with optional
            client-side filtering. Server-side ``status`` / ``project_id`` /
            etc. filters are silently ignored by TopKey's task index, so we
            fetch the full set then filter in Python."""
            try:
                # Fetch all tasks (auto-paginated by client.get_all). The
                # cap is 5000 — TopKey companies rarely have more open tasks
                # than that, and it caps memory.
                got = await c.get_all("/task", max_pages=200, max_items=5000)
                items = got.get("items", []) or []

                # Client-side filters (server ignores them).
                project_id = p.get("project_id")
                assigned_to = p.get("assigned_to")
                status = p.get("status")
                board_column_id = p.get("board_column_id")
                overdue_only = bool(p.get("overdue_only"))

                def _matches(t: dict) -> bool:
                    if project_id is not None and t.get("project") and \
                            (t["project"].get("id") if isinstance(t["project"], dict) else None) != int(project_id):
                        return False
                    if assigned_to is not None:
                        ids = [u.get("id") for u in (t.get("users") or []) if isinstance(u, dict)]
                        if int(assigned_to) not in ids:
                            return False
                    if status is not None and t.get("status") != status:
                        return False
                    if board_column_id is not None and t.get("board_column_id") != int(board_column_id):
                        return False
                    if overdue_only:
                        # Overdue = due_date in the past AND status != completed
                        due = t.get("due_date") or ""
                        if t.get("status") == "completed":
                            return False
                        if not due:
                            return False
                        try:
                            from datetime import datetime
                            due_dt = datetime.fromisoformat(due.replace("Z", "+00:00"))
                            if due_dt.timestamp() >= datetime.now(due_dt.tzinfo).timestamp():
                                return False
                        except (ValueError, TypeError):
                            return False
                    return True

                filtered = [t for t in items if _matches(t)]
                # Aggregate stats so the agent has counts without iterating
                from collections import Counter
                status_counts = Counter(t.get("status") for t in filtered)
                return self._maybe_spool(
                    "tasks",
                    filtered,
                    {
                        "fetched_total": len(items),
                        "match_total": len(filtered),
                        "status_counts": dict(status_counts),
                        "filters_applied": {
                            "project_id": project_id, "assigned_to": assigned_to,
                            "status": status, "board_column_id": board_column_id,
                            "overdue_only": overdue_only,
                        },
                    },
                )
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_list_tasks",
            (
                "TopKey: Hamma vazifalarni avtomatik to'la ro'yxatdan o'tkazib oladi (auto-pagination). "
                "Filterlar Python tomonida qo'llaniladi (server filterlari ishonchsiz). "
                "100+ qator natija avtomatik Excel faylga saqlanadi → preview + file_path qaytadi. "
                "Filter: project_id, assigned_to (user_id), status (completed/incomplete/in_progress), "
                "board_column_id, overdue_only (true: muddati o'tgan + bajarilmagan)."
            ),
            {"type": "object", "properties": {
                "project_id": {"type": "number"},
                "assigned_to": {"type": "number"},
                "status": {"type": "string", "description": "completed | incomplete | in_progress | review"},
                "board_column_id": {"type": "number"},
                "overdue_only": {"type": "boolean", "description": "Faqat muddati o'tgan + bajarilmaganlar"},
            }},
            list_tasks,
        ))

        async def get_task(p: dict) -> str:
            if "task_id" not in p:
                return self._err("task_id is required")
            try:
                return self._ok(await c.get(f"/task/{int(p['task_id'])}"))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_get_task",
            "TopKey: Vazifa to'liq ma'lumotlari (task_id bo'yicha).",
            {"type": "object", "required": ["task_id"], "properties": {
                "task_id": {"type": "number"},
            }},
            get_task,
        ))

        async def create_task(p: dict) -> str:
            for k in ("title", "project_id"):
                if k not in p:
                    return self._err(f"{k} is required")
            try:
                body: dict[str, Any] = {
                    "heading": p["title"],  # TopKey field name is `heading`
                    "title": p["title"],    # also accepted by some validators
                    "project_id": int(p["project_id"]),
                }
                if "description" in p:
                    body["description"] = p["description"]
                if "assigned_to" in p:
                    body["user_id"] = int(p["assigned_to"])
                if "due_date" in p:
                    body["due_date"] = p["due_date"]
                if "priority" in p:
                    body["priority"] = p["priority"]
                return self._ok(await c.post("/task", body))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_create_task",
            "TopKey: Yangi vazifa yaratish (title, project_id majburiy; assigned_to, due_date, priority ixtiyoriy).",
            {"type": "object", "required": ["title", "project_id"], "properties": {
                "title": {"type": "string"},
                "description": {"type": "string"},
                "project_id": {"type": "number"},
                "assigned_to": {"type": "number", "description": "user_id"},
                "due_date": {"type": "string", "description": "YYYY-MM-DD"},
                "priority": {"type": "string", "description": "low | medium | high"},
            }},
            create_task,
        ))

        async def update_task_status(p: dict) -> str:
            if "task_id" not in p:
                return self._err("task_id is required")
            if "status" not in p and "board_column_id" not in p:
                return self._err("either status or board_column_id is required")
            try:
                body: dict[str, Any] = {}
                if "status" in p:
                    body["status"] = p["status"]
                if "board_column_id" in p:
                    body["board_column_id"] = int(p["board_column_id"])
                return self._ok(await c.put(f"/task/{int(p['task_id'])}", body))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_update_task_status",
            "TopKey: Vazifa statusini o'zgartirish (status YOKI board_column_id orqali).",
            {"type": "object", "required": ["task_id"], "properties": {
                "task_id": {"type": "number"},
                "status": {"type": "string"},
                "board_column_id": {"type": "number"},
            }},
            update_task_status,
        ))

        async def list_subtasks(p: dict) -> str:
            if "task_id" not in p:
                return self._err("task_id is required")
            try:
                return self._ok(await c.get(f"/task/{int(p['task_id'])}/subtask"))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_list_subtasks",
            "TopKey: Vazifa ostidagi sub-vazifalar.",
            {"type": "object", "required": ["task_id"], "properties": {
                "task_id": {"type": "number"},
            }},
            list_subtasks,
        ))

        # ── WORK / TIME (2) ──────────────────────────────────

        async def log_time(p: dict) -> str:
            for k in ("task_id", "hours"):
                if k not in p:
                    return self._err(f"{k} is required")
            try:
                body: dict[str, Any] = {
                    "task_id": int(p["task_id"]),
                    "total_hours": float(p["hours"]),
                }
                if "date" in p:
                    body["date"] = p["date"]
                if "memo" in p:
                    body["memo"] = p["memo"]
                return self._ok(await c.post("/timelog", body))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_log_time",
            "TopKey: Vazifaga sarflangan vaqtni qayd qilish (task_id, hours; ixtiyoriy date, memo).",
            {"type": "object", "required": ["task_id", "hours"], "properties": {
                "task_id": {"type": "number"},
                "hours": {"type": "number"},
                "date": {"type": "string", "description": "YYYY-MM-DD"},
                "memo": {"type": "string"},
            }},
            log_time,
        ))

        async def list_my_timelogs(p: dict) -> str:
            try:
                params: dict[str, Any] = {}
                if "from_date" in p:
                    params["from_date"] = p["from_date"]
                if "to_date" in p:
                    params["to_date"] = p["to_date"]
                if "page" in p:
                    params["page"] = p["page"]
                if "per_page" in p:
                    params["per_page"] = p["per_page"]
                return self._ok(await c.get("/timelog/me", params or None))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_list_my_timelogs",
            "TopKey: Joriy foydalanuvchining vaqt qaydlari (from_date / to_date oralig'ida).",
            {"type": "object", "properties": {
                "from_date": {"type": "string"},
                "to_date": {"type": "string"},
                "page": {"type": "number"},
                "per_page": {"type": "number"},
            }},
            list_my_timelogs,
        ))

        # ── AUTH (3) ─────────────────────────────────────────

        async def login_tool(_p: dict) -> str:
            try:
                await c.login()
                return self._ok({"ok": True, "message": "Re-logged in successfully"})
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_login",
            "TopKey: Token qayta olish (majburiy login). Odatda kerak emas — 401 bo'lganda avto re-login ishlaydi.",
            {"type": "object", "properties": {}},
            login_tool,
        ))

        async def get_current_user(_p: dict) -> str:
            try:
                return self._ok(await c.get("/auth/me"))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_get_current_user",
            "TopKey: Hozirda qaysi foydalanuvchi (admin) sifatida bot ishlayotgani.",
            {"type": "object", "properties": {}},
            get_current_user,
        ))

        async def list_users(p: dict) -> str:
            try:
                params: dict[str, Any] = {}
                if "page" in p:
                    params["page"] = p["page"]
                if "per_page" in p:
                    params["per_page"] = p["per_page"]
                return self._ok(await c.get("/user", params or None))
            except Exception as e:
                return self._err(str(e))
        tools.append(ToolDef(
            "topkey_list_users",
            "TopKey: Tizim foydalanuvchilari ro'yxati (admin only).",
            {"type": "object", "properties": {
                "page": {"type": "number"},
                "per_page": {"type": "number"},
            }},
            list_users,
        ))

        return tools
